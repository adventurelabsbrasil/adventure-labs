"""Inspeciona arquivo no Drive por mimeType."""

from __future__ import annotations

import io
import re
from typing import Any

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

MIME_DOC = "application/vnd.google-apps.document"
MIME_SHEET = "application/vnd.google-apps.spreadsheet"
MIME_SLIDES = "application/vnd.google-apps.presentation"
MIME_CSV = "text/csv"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_XLS = "application/vnd.ms-excel"


def _extract_slide_text(slide: dict[str, Any]) -> list[str]:
    out: list[str] = []
    for el in slide.get("pageElements", []) or []:
        shape = el.get("shape")
        if not shape:
            continue
        tc = shape.get("text", {})
        for te in tc.get("textElements", []) or []:
            tr = te.get("textRun")
            if tr and tr.get("content"):
                t = str(tr["content"]).strip()
                if t:
                    out.append(t[:500])
    return out


def _sheet_range(tab: str, max_rows: int) -> str:
    if re.search(r"[^A-Za-z0-9_]", tab):
        safe = tab.replace("'", "''")
        return f"'{safe}'!A1:Z{max_rows}"
    return f"{tab}!A1:Z{max_rows}"


def _build_summary(structure: dict[str, Any], flags: list[str]) -> str:
    name = structure.get("name") or "Arquivo"
    t = structure.get("type") or ""
    extra = f" {len(flags)} alerta(s)." if flags else ""
    if MIME_SLIDES in t:
        n = structure.get("slideCount", 0)
        return f"Apresentacao '{name}' com {n} slide(s).{extra}"
    if structure.get("tabs"):
        return f"Planilha/tabular '{name}' com {len(structure['tabs'])} aba(s) amostradas.{extra}"
    if structure.get("stmtTransactionBlocks") is not None:
        tr = structure["stmtTransactionBlocks"]
        return f"OFX '{name}' com ~{tr} bloco(s) STMTTRN.{extra}"
    return f"'{name}' ({t or 'tipo misto'}).{extra}"


def inspect_file(
    creds: Credentials,
    file_id: str,
    *,
    max_rows: int = 50,
    max_sheets: int = 10,
    max_slides: int = 40,
) -> dict[str, Any]:
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    meta = (
        drive.files()
        .get(
            fileId=file_id,
            fields="id,name,mimeType,size",
            supportsAllDrives=True,
        )
        .execute()
    )
    mime = meta.get("mimeType") or ""
    name = meta.get("name") or ""
    quality_flags: list[str] = []
    suggested: list[str] = []
    structure: dict[str, Any] = {"type": mime, "name": name}

    if mime == MIME_DOC:
        data = drive.files().export(fileId=file_id, mimeType="text/plain").execute()
        text = data.decode("utf-8", errors="replace") if isinstance(data, bytes) else str(data)
        structure["charCount"] = len(text)
        structure["sampleLines"] = text.splitlines()[:25]
        if len(text) > 80_000:
            quality_flags.append("Documento muito longo; amostra parcial.")
        suggested.append("Sumario no topo e headings consistentes (H1/H2).")

    elif mime == MIME_SHEET:
        sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
        ss = (
            sheets.spreadsheets()
            .get(
                spreadsheetId=file_id,
                fields="properties.title,sheets.properties.title",
            )
            .execute()
        )
        titles = [s["properties"]["title"] for s in ss.get("sheets", [])][
            :max_sheets
        ]
        structure["tabs"] = titles
        tab_samples: dict[str, list[list[str]]] = {}
        for t in titles:
            try:
                res = (
                    sheets.spreadsheets()
                    .values()
                    .get(
                        spreadsheetId=file_id,
                        range=_sheet_range(t, max_rows),
                    )
                    .execute()
                )
                rows = res.get("values") or []
            except Exception:
                rows = []
            tab_samples[t] = [[str(c)[:200] for c in r[:26]] for r in rows[:max_rows]]
            if rows:
                hdr = [str(x).strip() for x in rows[0]]
                non_empty = [h for h in hdr if h]
                if len(non_empty) != len(set(non_empty)):
                    quality_flags.append(
                        f"Aba '{t}': possiveis headers duplicados ou colunas sem nome."
                    )
        structure["tabRowCounts"] = {k: len(v) for k, v in tab_samples.items()}
        structure["tabSamples"] = {k: v[:5] for k, v in tab_samples.items()}
        suggested.append(
            "Padronizar nomes de abas; uma linha de header; evitar celulas mescladas no cabecalho."
        )

    elif mime == MIME_SLIDES:
        slides = build("slides", "v1", credentials=creds, cache_discovery=False)
        p = slides.presentations().get(presentationId=file_id).execute()
        slides_out: list[dict[str, Any]] = []
        for slide in (p.get("slides") or [])[:max_slides]:
            oid = slide.get("objectId", "")
            texts = _extract_slide_text(slide)
            slides_out.append(
                {
                    "objectId": oid,
                    "textSnippets": texts[:12],
                    "snippetCount": len(texts),
                }
            )
        structure["slideCount"] = len(p.get("slides") or [])
        structure["slidesSample"] = slides_out
        if all(s["snippetCount"] == 0 for s in slides_out):
            quality_flags.append(
                "Pouco texto em shapes; conteudo pode estar em imagens ou tabelas."
            )
        suggested.append(
            "Um titulo forte por slide; detalhes nas notas; limitar texto por slide."
        )

    else:
        req = drive.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        raw = buf.getvalue()
        nl = name.lower()

        if mime == MIME_CSV or nl.endswith(".csv"):
            text = raw.decode("utf-8", errors="replace")[:300_000]
            lines = text.splitlines()
            structure["lineCount"] = len(lines)
            structure["sampleLines"] = lines[: min(25, len(lines))]
            suggested.append("UTF-8, cabecalho na linha 1, separador consistente.")

        elif mime == MIME_XLSX or nl.endswith(".xlsx"):
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sn = wb.sheetnames[:max_sheets]
            structure["tabs"] = sn
            samples: dict[str, list[list[str]]] = {}
            for sname in sn:
                ws = wb[sname]
                rows: list[list[str]] = []
                for i, row in enumerate(
                    ws.iter_rows(max_row=max_rows, max_col=26, values_only=True)
                ):
                    rows.append([("" if c is None else str(c))[:200] for c in row])
                samples[sname] = rows
                if rows and not any(str(x).strip() for x in rows[0]):
                    quality_flags.append(f"Aba '{sname}': header parece vazio.")
            wb.close()
            structure["tabSamples"] = {k: v[:5] for k, v in samples.items()}

        elif mime == MIME_XLS or nl.endswith(".xls"):
            import xlrd

            wb = xlrd.open_workbook(file_contents=raw)
            sn = wb.sheet_names()[:max_sheets]
            structure["tabs"] = sn
            samples = {}
            for sname in sn:
                sh = wb.sheet_by_name(sname)
                rows = []
                for rx in range(min(max_rows, sh.nrows)):
                    rows.append(
                        [
                            str(sh.cell_value(rx, cx))[:200]
                            for cx in range(min(26, sh.ncols))
                        ]
                    )
                samples[sname] = rows
            structure["tabSamples"] = {k: v[:5] for k, v in samples.items()}
            suggested.append(
                "Preferir migrar .xls para .xlsx ou Google Sheets (formato legado)."
            )

        elif "ofx" in mime.lower() or nl.endswith(".ofx"):
            text = raw.decode("utf-8", errors="replace")[:800_000]
            tr = text.upper().count("<STMTTRN>")
            structure["stmtTransactionBlocks"] = tr
            structure["ofxHead"] = text[:1500]
            if tr == 0:
                quality_flags.append("Nenhum STMTTRN encontrado.")
            suggested.append(
                "Conferir datas, contas e saldo final; reconciliar com extrato."
            )

        else:
            structure["sizeBytes"] = len(raw)
            quality_flags.append(f"Tipo nao especializado ({mime}).")
            suggested.append(
                "Converter para CSV/Sheets ou exportar texto para inspecao mais profunda."
            )

    summary = _build_summary(structure, quality_flags)
    if not suggested:
        suggested = ["Revisar nome do arquivo e pasta no Drive para padrao da empresa."]
    return {
        "fileId": file_id,
        "name": name,
        "mimeType": mime,
        "summary": summary.strip(),
        "detectedStructure": structure,
        "qualityFlags": quality_flags,
        "suggestedImprovements": suggested[:15],
    }
